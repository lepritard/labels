#!/usr/bin/env python3
"""
parse_packing_slip.py  —  Evolved packing slip parser (v2, v1.48)

Reads a New Arch ENTERPRISE packing slip .xlsx (Sheet2) and returns
structured JSON describing every label group to be printed.

Supports:
  - Standard carton rows  ("Carton A  (PART :N PCS / CARTON)")
  - Co-packed secondary part numbers (col A blank, col B = secondary PN)
  - Pallet rows           ("Pallet A  (PART: N PCS / PALLET)")
  - Wooden case rows      ("Wooden case A  (PART: N PCS / WOODEN CASE)")
  - Mixed shipments (cartons + pallets + wooden cases in one file)
  - "PALLET N" annotation in col G (carton groups assigned to pallets)
  - "T013 PREPPED" flag in col G
  - TBCTW- prefix typo correction
  - Bare numeric part numbers via KNOWN_PARTS map
  - EXCEPTIONS dict for co-packed inner-box label counts
  - Revision stripping for EXCEPTIONS lookup (keeps revision in output)

Output JSON schema (one object per file):
{
  "meta": {
    "po":           "PO22643",
    "date":         "APR.,15, 2026",
    "container_no": "CSNU6534955",
    "seal":         "OOLKBA5958"
  },
  "label_groups": [
    {
      "customer":        "TurboChef",
      "part_number":     "TBC-ENC-1726-E",
      "base_part":       "ENC-1726",          // revision stripped
      "revision":        "E",                 // None if no revision
      "pcs_per_box":     4,
      "num_labels":      41,
      "container_type":  "carton",            // carton | pallet | wooden_case
      "labels_per_unit": 1,                   // 1 for cartons, 4 for pallets/wooden cases
      "total_labels":    41,                  // num_labels * labels_per_unit
      "flags":           [],                  // ["t013_prepped", "co_packed_secondary", ...]
      "pallet_group":    null                 // "PALLET 1" or null
    },
    ...
  ]
}
"""

import re
import json
import sys
import openpyxl
from collections import defaultdict
from pathlib import Path


# ---------------------------------------------------------------------------
# Configuration tables — edit these as new customers / exceptions are found
# ---------------------------------------------------------------------------

CUSTOMER_MAP = {
    "TBC":    "TurboChef",
    "TAY":    "Taylor Company",
    "APS":    "Avtron Power Solutions",
    "AII":    "Appliance Innovation",
    "DWG":    "DAWGS",
    "JJP":    "Jersey Jack Pinball",
    "MCS":    "Middleby Coffee Solutions",
    "BPI":    "Berryman Products",
    "CPI":    "Commercial Plastics",
    "SCI":    "Stericycle",
    "FRE":    "TurboChef",
    "IB":    "DAWGS",
    "NGC":    "TurboChef",
    "HCW":    "TurboChef",
    "HHD":    "TurboChef",       # HHD prefix is TurboChef product line
    "HCT":    "TurboChef",
    "HHS":    "TurboChef",
    "ECO":    "TurboChef",
    "ECS":    "TurboChef",
    "ENC":    "TurboChef",
    "SKU":    "Unknown",
    "TBCTW":  "TurboChef",       # typo variant — handled specially below
}

# Parts identified solely by their full part number (no alphabetic prefix pattern)
KNOWN_PARTS = {
    "400009838": {"customer": "LAB2FAB", "base_part": "400009838", "revision": None},
}

# Co-packed inner-box rules.
# Prefixes whose labels should omit the customer name entirely.
# These parts ship to intermediary companies; the customer field is
# set correctly for internal reference but suppressed on the printed label.
OMIT_CUSTOMER_PREFIXES = {"FRE"}

# Key = base part number WITHOUT revision letter (e.g. "ENC-1833", not "TBC-ENC-1833-B")
# labels_per_carton: how many inner boxes (= labels) exist inside each outer carton
# pcs_per_label: pieces in each inner box
EXCEPTIONS = {
    "ENC-1725": {"labels_per_carton": 2, "pcs_per_label": 2},
    "ENC-1833": {"labels_per_carton": 2, "pcs_per_label": 3},
    "ENC-1724": {"labels_per_carton": 2, "pcs_per_label": 6},
    "HHD-8301": {"labels_per_carton": 3, "pcs_per_label": 2},
}

# Labels printed per physical unit
LABELS_PER_UNIT = {
    "carton":      1,
    "pallet":      4,
    "wooden_case": 4,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _s(cell):
    """Return cell value as stripped string, or empty string."""
    return str(cell).strip() if cell is not None else ""


def extract_meta(rows):
    meta = {"po": None, "date": None, "container_no": None, "seal": None}
    for row in rows[:12]:
        for cell in row:
            v = _s(cell)
            if not v:
                continue
            m = re.search(r"No\.?:\s*(\S+)", v)
            if m and meta["po"] is None:
                meta["po"] = m.group(1)
            m = re.search(r"Date:\s*(.+)", v)
            if m and meta["date"] is None:
                meta["date"] = m.group(1).strip()
            m = re.search(r"CONTAINER\s*NO[:\s]+(\S+)", v, re.IGNORECASE)
            if m and meta["container_no"] is None:
                meta["container_no"] = m.group(1).strip()
            m = re.search(r"SEAL\s*NUMBER\s*[:\s]+(\S+)", v, re.IGNORECASE)
            if m and meta["seal"] is None:
                meta["seal"] = m.group(1).strip()
    return meta


def find_header_row(rows):
    for i, row in enumerate(rows):
        vals = [_s(v).upper() for v in row]
        if any("DESCRIPTION" in v for v in vals) and any("QUANTITY" in v for v in vals):
            return i
    return None


def find_data_end(rows, start):
    for i, row in enumerate(rows[start:], start):
        v = _s(row[0]).upper()
        if v.startswith("TOTAL") or v.startswith("SAY TOTAL"):
            return i
    return len(rows)


def detect_container_type(col_a):
    a = col_a.lower()
    if a.startswith("pallet"):
        return "pallet"
    if a.startswith("wooden case"):
        return "wooden_case"
    if a.startswith("carton"):
        return "carton"
    return "carton"  # default


def parse_qty(val):
    """Extract integer from a cell that may be int, float, or string like '10 PALLETS'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    m = re.search(r"(\d[\d,]*)", str(val).replace(",", ""))
    return int(m.group(1)) if m else None


def pcs_from_label_text(part_number, label_text):
    """
    Extract pieces-per-box from the verbose col A label text.
    Tries to match the part number then find the PCS figure after it.
    Falls back to any single PCS figure in the text.
    """
    if not label_text:
        return None
    esc = re.escape(part_number)
    # e.g. "TBC-ENC-1726-E :4 PCS / CARTON"
    m = re.search(esc + r"[^0-9]{0,80}?(\d[\d,]*)\s*PCS", label_text, re.IGNORECASE)
    if m:
        return int(m.group(1).replace(",", ""))
    # Fallback: any PCS figures
    nums = re.findall(r"(\d[\d,]*)\s*PCS", label_text, re.IGNORECASE)
    if len(nums) == 1:
        return int(nums[0].replace(",", ""))
    return None


def normalise_part(raw_part):
    """
    Given a raw part number string from col B, return:
      (display_part, customer, base_for_exceptions, revision, omit_customer)

    Handles:
      - Standard "PREFIX-REST-REV" format
      - TBCTW typo → TBC prefix, display as "TW-REST-REV"
      - Bare numeric parts via KNOWN_PARTS
      - TBC sub-prefixes (ENC, HHD, etc.) mapped to TurboChef directly
      - Trailing " REV<N>" / "-REV<N>" suffix (e.g. "TBC-CT-104546 REV2")
      - T-series tooling suffix (e.g. "TBC-ECO-9447-B-T013" → base "ECO-9447-T013", rev "B")
    """
    raw = raw_part.strip()

    # Pre-strip trailing " REV<N>" or "-REV<N>" suffix (case-insensitive).
    # e.g. "TBC-CT-104546 REV2" → raw="TBC-CT-104546", _pre_rev="2"
    # Captured revision may be overridden by later dash-suffix logic, but this
    # ensures the token is removed from the part number string first.
    _pre_rev = None
    _rev_sfx = re.search(r'[\s-]REV\s*([A-Z0-9]+)$', raw, re.IGNORECASE)
    if _rev_sfx:
        _pre_rev = _rev_sfx.group(1)
        raw = raw[:_rev_sfx.start()].strip()

    # Bare numeric
    if re.match(r"^\d+$", raw):
        if raw in KNOWN_PARTS:
            kp = KNOWN_PARTS[raw]
            return raw, kp["customer"], kp["base_part"], kp["revision"],\
                   False  # known numeric parts never omit customer
        return raw, "Unknown", raw, None, False

    # TBCTW typo: e.g. TBCTW-I1-9988-B → display "TW-I1-9988-B", customer TurboChef
    if raw.upper().startswith("TBCTW-"):
        display = raw[3:]          # strip "TBC" → "TW-I1-9988-B"
        rest = display             # e.g. "TW-I1-9988-B"
        # strip revision from base key (single letter OR letter+digits, e.g. B1)
        rev_m = re.search(r"-([A-Z]\d+|[A-Z])$", rest)
        revision = rev_m.group(1) if rev_m else (_pre_rev or None)
        base = rest[:-(len(revision) + 1)] if rev_m else rest
        return display, "TurboChef", base, revision, False  # TBCTW never omits customer

    # Standard: split on first dash to get prefix
    # Extract leading alphabetic characters as the customer prefix
    prefix_m = re.match(r'^([A-Za-z]+)', raw)
    prefix = prefix_m.group(1).upper() if prefix_m else raw.split('-')[0].upper()

    # T-series tooling suffix: PREFIX-BASE-REV-Txxxx
    # e.g. TBC-ECO-9447-B-T013 → display "ECO-9447-T013", revision "B"
    # Detect pattern: after stripping any leading PREFIX-, the rest matches BASE-LETTER-Tdddd
    _t_series = re.search(r'^(.+)-([A-Z])-(T\d+)$', raw, re.IGNORECASE)
    if _t_series:
        # Rebuild: strip leading customer prefix from group(1) for display
        _t_base_raw = _t_series.group(1)   # e.g. "TBC-ECO-9447"
        _t_rev      = _t_series.group(2)   # e.g. "B"
        _t_sfx      = _t_series.group(3)   # e.g. "T013"
        # Strip customer prefix for display/exc_key
        _t_strip = re.match(r'^[A-Za-z]{2,4}-(.+)$', _t_base_raw)
        _t_base_clean = _t_strip.group(1) if _t_strip else _t_base_raw
        display   = f"{_t_base_clean}-{_t_sfx}"   # ECO-9447-T013
        exc_key   = display                        # used for EXCEPTIONS lookup
        revision  = _t_rev.upper()
        prefix_m2 = re.match(r'^([A-Za-z]+)-', raw)
        prefix2   = prefix_m2.group(1).upper() if prefix_m2 else ""
        customer  = CUSTOMER_MAP.get(prefix2, f"Unknown ({prefix2})")
        return display, customer, exc_key, revision,\
               (prefix2 in OMIT_CUSTOMER_PREFIXES)

    # Revision is last token if it's a single capital letter OR letter+digits (e.g. B1, C2)
    rev_m = re.search(r"-([A-Z]\d+|[A-Z])$", raw)
    revision = rev_m.group(1) if rev_m else (_pre_rev or None)
    base = raw[:-(len(revision) + 1)] if rev_m else raw  # strip "-REV" suffix for EXCEPTIONS key

    # Strip leading PREFIX- from base_part for barcode display.
    # For TBC parts the sub-prefix (ENC-, HHD-, etc.) is the "real" part number.
    # For all other customers (APS, TAY, MCS, BPI, etc.) the prefix is just the
    # customer code and should not appear on the label.
    exc_key = base
    if exc_key.upper().startswith("TBC-"):
        exc_key = exc_key[4:]
    else:
        # Strip any known 2-4 char alpha prefix followed by a dash.
        # Exception: prefixes in OMIT_CUSTOMER_PREFIXES keep their prefix in
        # the barcode because the prefix is part of the part identity, not
        # just a customer code (e.g. FRE-1160 barcodes as "FRE-1160", not "1160").
        strip_m = re.match(r'^([A-Za-z]{2,4})-(.+)$', exc_key)
        if (strip_m
                and strip_m.group(1).upper() in CUSTOMER_MAP
                and strip_m.group(1).upper() not in OMIT_CUSTOMER_PREFIXES):
            exc_key = strip_m.group(2)

    customer = CUSTOMER_MAP.get(prefix, f"Unknown ({prefix})")

    return raw, customer, exc_key, revision,\
           (prefix.upper() in OMIT_CUSTOMER_PREFIXES)


# ---------------------------------------------------------------------------
# Core parser
# ---------------------------------------------------------------------------

def parse_packing_slip(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    # Try Sheet2 first (New Arch default); fall back to first sheet
    ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    meta = extract_meta(rows)
    header_row = find_header_row(rows)
    if header_row is None:
        raise ValueError(f"Could not find header row in {filepath}")

    data_end = find_data_end(rows, header_row + 1)
    label_groups = []

    pending_secondary = []   # accumulates secondary PNs until next primary row

    def flush_secondary(primary_group, secondary_pns, primary_col_a=""):
        """Attach secondary (co-packed) PNs to the primary group's carton count."""
        for sec_pn in secondary_pns:
            display, customer, exc_key, revision, omit_cust = normalise_part(sec_pn)
            exc = EXCEPTIONS.get(exc_key, {})
            lpc = exc.get("labels_per_carton", 1)
            ppl = exc.get("pcs_per_label", None)
            # Fallback: try to read pcs for this secondary from
            # the primary row's col A description text (the supplier
            # often lists all co-packed quantities in one cell).
            if ppl is None and primary_col_a:
                ppl = pcs_from_label_text(sec_pn, primary_col_a)
            num_outer = primary_group["num_labels"]
            total = num_outer * lpc

            label_groups.append({
                "customer":        customer,
                "part_number":     display,
                "base_part":       exc_key,
                "revision":        revision,
                "pcs_per_box":     ppl,
                "num_labels":      total,
                "container_type":  "carton",
                "labels_per_unit": 1,
                "total_labels":    total,
                "flags":           ["co_packed_secondary"],
                "omit_customer":   omit_cust,
                "co_packed_parent": primary_group["base_part"],
                "pallet_group":    primary_group.get("pallet_group"),
            })

    last_primary = None
    last_primary_col_a = ""

    for row in rows[header_row + 1:data_end]:
        col_a = _s(row[0])
        col_b = _s(row[1])
        col_c = row[2] if len(row) > 2 else None
        col_g = _s(row[6]) if len(row) > 6 else ""

        # Skip blank rows
        if not col_a and not col_b:
            continue

        # Secondary (co-packed) row: col A blank, col B has part number
        if not col_a and col_b:
            pending_secondary.append(col_b)
            continue

        # Primary row — first flush any pending secondary PNs onto the PREVIOUS primary
        if pending_secondary and last_primary is not None:
            flush_secondary(last_primary, pending_secondary, last_primary_col_a)
            pending_secondary = []

        # Parse primary row
        ctype = detect_container_type(col_a)
        qty_units = parse_qty(col_c)
        if qty_units is None:
            continue

        display, customer, exc_key, revision, omit_cust = normalise_part(col_b)
        pcs = pcs_from_label_text(col_b, col_a)

        # Flags
        flags = []
        pallet_group = None
        if "T013" in col_g.upper():
            flags.append("t013_prepped")
        pallet_m = re.search(r"PALLET\s*(\d+)", col_g, re.IGNORECASE)
        if pallet_m:
            pallet_group = f"PALLET {pallet_m.group(1)}"

        lpu = LABELS_PER_UNIT.get(ctype, 1)
        total_labels = qty_units * lpu

        group = {
            "customer":        customer,
            "part_number":     display,
            "base_part":       exc_key,
            "revision":        revision,
            "pcs_per_box":     pcs,
            "num_labels":      qty_units,
            "container_type":  ctype,
            "labels_per_unit": lpu,
            "total_labels":    total_labels,
            "flags":           flags,
            "omit_customer":   omit_cust,
            "pallet_group":    pallet_group,
        }

        label_groups.append(group)
        last_primary = group
        last_primary_col_a = col_a

    # Flush any trailing secondary PNs
    if pending_secondary and last_primary is not None:
        flush_secondary(last_primary, pending_secondary, last_primary_col_a)


    # ── Post-process step 0: reconcile inverted co-pack remainder blocks ──
    # The supplier occasionally lists the remainder box of a co-packed carton
    # with the part order inverted (secondary part first, primary part second).
    # Detection: a primary group at index i whose base_part previously appeared
    # as a co_packed_secondary of an earlier group, AND whose own co_packed
    # secondary's base_part is that earlier primary — i.e., both part numbers
    # appear in both positions, just swapped.
    # The earlier group is found by scanning indices 0..i-1 (never i itself),
    # so the two ENC-1726 entries are never confused with each other.
    # Action:
    #   - Add the inverted group's num_labels to the original primary's count.
    #   - If the inverted group's pcs differs from the standard secondary pcs,
    #     record a nonstd_remainders entry on the original secondary group.
    #   - Remove both inverted groups from label_groups.

    to_remove = set()
    for i, inv_primary in enumerate(label_groups):
        if i in to_remove:
            continue
        if "co_packed_secondary" in inv_primary.get("flags", []):
            continue
        if "nonstd_remainder" in inv_primary.get("flags", []):
            continue
        if inv_primary["container_type"] != "carton":
            continue

        # Find inv_primary's co-packed secondary (must come after i)
        inv_secondary = None
        for j in range(i + 1, len(label_groups)):
            g2 = label_groups[j]
            if ("co_packed_secondary" in g2.get("flags", [])
                    and g2.get("co_packed_parent") == inv_primary["base_part"]):
                inv_secondary = (j, g2)
                break
        if inv_secondary is None:
            continue
        j_sec, inv_sec_group = inv_secondary

        # Find the original primary: a NON-secondary group at index < i
        # whose base_part == inv_sec_group["base_part"]
        orig_primary = None
        orig_primary_idx = None
        for k in range(i):
            g2 = label_groups[k]
            if (g2["base_part"] == inv_sec_group["base_part"]
                    and "co_packed_secondary" not in g2.get("flags", [])
                    and "nonstd_remainder" not in g2.get("flags", [])
                    and g2["container_type"] == "carton"):
                orig_primary = g2
                orig_primary_idx = k
                break
        if orig_primary is None:
            continue

        # Find the original secondary: a co_packed_secondary whose
        # co_packed_parent == orig_primary["base_part"]
        # AND whose base_part == inv_primary["base_part"]
        orig_sec = None
        for k in range(len(label_groups)):
            g2 = label_groups[k]
            if (k != i
                    and "co_packed_secondary" in g2.get("flags", [])
                    and g2.get("co_packed_parent") == orig_primary["base_part"]
                    and g2["base_part"] == inv_primary["base_part"]):
                orig_sec = g2
                break
        if orig_sec is None:
            continue

        # All checks passed — reconcile.
        n_inv = inv_primary["num_labels"]
        new_total = orig_primary["num_labels"] + n_inv

        # 1. Grow the original primary's count
        orig_primary["num_labels"]   = new_total
        orig_primary["total_labels"] = new_total

        # 2. Update original secondary total to include the remainder boxes
        orig_sec["num_labels"]   = orig_sec["num_labels"] + n_inv
        orig_sec["total_labels"] = orig_sec["total_labels"] + n_inv

        # 3. If the inverted primary's pcs differs from the original secondary's
        #    standard pcs, record a nonstd_remainder on the original secondary.
        #    nonstd_box_num = secondary's final total (computed above in step 2).
        #    nonstd_copies  = 5 (NON-STD labels always print 5 copies).
        if inv_primary["pcs_per_box"] != orig_sec["pcs_per_box"]:
            orig_sec.setdefault("nonstd_remainders", []).append({
                "nonstd_box_num": orig_sec["num_labels"],
                "nonstd_pcs":     inv_primary["pcs_per_box"],
                "nonstd_copies":  5,
            })
            if "has_nonstd_remainder" not in orig_sec["flags"]:
                orig_sec["flags"].append("has_nonstd_remainder")

        # 4. Remove the two inverted groups
        to_remove.add(i)
        to_remove.add(j_sec)

    label_groups = [g for k, g in enumerate(label_groups) if k not in to_remove]


    # ── Post-process step 1: consolidate scattered same-part rows ─────────
    # Parts that appear in multiple non-consecutive runs (e.g. same PN spread
    # across pallets in the packing slip) are merged into a single entry.
    # Merge key = (base_part, container_type, pcs_per_box).
    # co_packed_secondary rows are NOT merged (they depend on their parent count).
    # Rows with different pallet_group values are kept separate.
    from collections import OrderedDict
    merged: dict = OrderedDict()
    unmerged = []
    for g in label_groups:
        is_secondary = "co_packed_secondary" in g.get("flags", [])
        # Key: part + type + pcs. pallet_group kept separate if set.
        key = (g["base_part"], g["container_type"], g["pcs_per_box"], g.get("pallet_group"))
        if is_secondary or g["container_type"] != "carton":
            # Pallets, wooden cases, and co-packed secondaries are never merged
            unmerged.append(("__unmerged__", g))
            continue
        if key in merged:
            merged[key]["num_labels"] += g["num_labels"]
            merged[key]["total_labels"] += g["total_labels"]
            # Union flags (excluding nonstd_remainder from sub-rows)
            for fl in g.get("flags", []):
                if fl not in merged[key]["flags"]:
                    merged[key]["flags"].append(fl)
            # Keep the first pallet_group seen (or override if previously None)
            if merged[key]["pallet_group"] is None and g.get("pallet_group"):
                merged[key]["pallet_group"] = g["pallet_group"]
        else:
            merged[key] = dict(g)  # copy so we don't mutate the original

    # Rebuild label_groups: merged cartons in order of first appearance, then unmerged
    # Preserve original order by interleaving based on OrderedDict insertion order.
    # Non-carton rows stay in their original relative order.
    # Strategy: walk original list; on first encounter of a key emit the merged entry,
    # on subsequent encounters skip. Non-carton rows emitted in place.
    seen_keys: set = set()
    label_groups_new = []
    unmerged_iter = iter(unmerged)
    for g in label_groups:
        is_secondary = "co_packed_secondary" in g.get("flags", [])
        key = (g["base_part"], g["container_type"], g["pcs_per_box"], g.get("pallet_group"))
        if is_secondary or g["container_type"] != "carton":
            label_groups_new.append(next(unmerged_iter)[1])
        else:
            if key not in seen_keys:
                seen_keys.add(key)
                label_groups_new.append(merged[key])
            # else: duplicate — already emitted, skip
    label_groups = label_groups_new

    # ── Post-process step 2: detect remainder (non-standard) boxes ─────────
    # After merging, check for consecutive same-part carton rows where the
    # second has fewer boxes — that's the non-standard remainder.
    # Note: after consolidation, genuine nonstd remainders only occur when
    # the pcs_per_box differs between rows of the same base_part (different
    # box sizes). Same-pcs rows are fully merged above.
    for i in range(1, len(label_groups)):
        cur = label_groups[i]
        prev = label_groups[i - 1]
        _same_base   = cur["base_part"] == prev["base_part"]
        _both_carton = cur["container_type"] == "carton" and prev["container_type"] == "carton"
        _not_flagged = ("nonstd_remainder" not in cur["flags"]
                        and "co_packed_secondary" not in cur["flags"]
                        and "nonstd_remainder" not in prev["flags"]
                        and "co_packed_secondary" not in prev["flags"])
        # Trigger when: fewer boxes (classic non-std remainder) OR
        # different pcs_per_box with same base_part (qty-variant non-std box).
        _fewer_boxes = cur["num_labels"] < prev["num_labels"]
        _diff_pcs    = cur["pcs_per_box"] != prev["pcs_per_box"]
        if _same_base and _both_carton and _not_flagged and (_fewer_boxes or _diff_pcs):
            cur["flags"].append("nonstd_remainder")
            # nonstd_box_num = first serial of this sub-group (prev boxes + 1)
            cur["nonstd_box_num"] = prev["num_labels"] + 1
            cur["nonstd_copies"] = 5
            # Merge total into the parent so total_labels is accurate
            prev["total_boxes"] = prev["num_labels"] + cur["num_labels"]

    return {"meta": meta, "label_groups": label_groups}


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python parse_packing_slip.py <file.xlsx> [file2.xlsx ...]")
        sys.exit(1)

    results = {}
    for path in sys.argv[1:]:
        try:
            data = parse_packing_slip(path)
            results[Path(path).name] = data
        except Exception as e:
            results[Path(path).name] = {"error": str(e)}

    print(json.dumps(results, indent=2))


if __name__ == "__main__":
    main()
